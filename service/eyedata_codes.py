import argparse
import re
import xml.etree.ElementTree as ET
from datetime import datetime
from pathlib import Path
from xml.etree.ElementTree import Element

from openpyxl import Workbook

from utils.config_manager import ConfigManager

XML_ENCODING = "cp932"
CODE_PATTERN = re.compile(r"^(\d+)(.*)$")
EXCLUDED_TAGS = frozenset({"SumStaff", "Kind4"})
INVALID_SHEET_NAME_CHARS = re.compile(r"[:\\/?*\[\]]")
SHEET_NAME_MAX_LENGTH = 31


def load_root(xml_path: Path) -> Element:
    return ET.fromstring(xml_path.read_bytes().decode(XML_ENCODING))


def iter_code_entries(element: Element) -> list[tuple[str, str, str]]:
    """(要素名, コード, 名称) を再帰的に収集する。カンマ区切りのコードは分割する。"""
    entries: list[tuple[str, str, str]] = []
    for child in element:
        raw_code = (child.findtext("Code") or "").strip()
        if raw_code and child.tag not in EXCLUDED_TAGS:
            name = (child.findtext("Name") or "").strip()
            for code in raw_code.split(","):
                if code.strip():
                    entries.append((child.tag, code.strip(), name))
        entries.extend(iter_code_entries(child))
    return entries


def aggregate_by_tag_and_number(
    entries: list[tuple[str, str, str]],
) -> dict[tuple[str, int], tuple[set[str], set[str]]]:
    """(要素名, 数値部分) をキーに (名称, 接尾辞) を集約する。"""
    aggregated: dict[tuple[str, int], tuple[set[str], set[str]]] = {}
    for tag, code, name in entries:
        matched = CODE_PATTERN.match(code)
        if matched is None:
            continue
        names, suffixes = aggregated.setdefault(
            (tag, int(matched.group(1))), (set(), set())
        )
        if name:
            names.add(name)
        if matched.group(2):
            suffixes.add(matched.group(2))
    return aggregated


def find_free_ranges(used: set[int], maximum: int) -> list[tuple[int, int]]:
    """1 から maximum までの未使用番号を連続範囲にまとめる。"""
    ranges: list[tuple[int, int]] = []
    start: int | None = None
    for number in range(1, maximum + 1):
        if number in used:
            if start is not None:
                ranges.append((start, number - 1))
                start = None
        elif start is None:
            start = number
    if start is not None:
        ranges.append((start, maximum))
    return ranges


def sanitize_sheet_name(tag: str) -> str:
    """Excel のシート名制約（使用不可文字・31文字以内）に収める。"""
    return INVALID_SHEET_NAME_CHARS.sub("_", tag)[:SHEET_NAME_MAX_LENGTH]


def write_excel(
    output_dir: Path, aggregated: dict[tuple[str, int], tuple[set[str], set[str]]]
) -> Path:
    """要素名ごとにシートを分けて Excel ファイルに出力する。"""
    output_dir.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    default_sheet = workbook.active
    if default_sheet is not None:
        workbook.remove(default_sheet)

    for tag, number in sorted(aggregated):
        sheet_name = sanitize_sheet_name(tag)
        sheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else None
        if sheet is None:
            sheet = workbook.create_sheet(sheet_name)
            sheet.append(["コード", "名称", "接尾辞"])
        names, suffixes = aggregated[(tag, number)]
        sheet.append([number, " / ".join(sorted(names)), " / ".join(sorted(suffixes))])

    timestamp = datetime.now().astimezone().strftime("%Y%m%d%H%M%S")
    excel_path = output_dir / f"eyedata_codes_{timestamp}.xlsx"
    workbook.save(excel_path)
    return excel_path


def format_range(start: int, end: int) -> str:
    return str(start) if start == end else f"{start}-{end}"


def main() -> None:
    project_root = Path(__file__).resolve().parents[1]
    parser = argparse.ArgumentParser(
        description="EyeData.xml のコード一覧と空き番号を出力する"
    )
    parser.add_argument(
        "xml",
        nargs="?",
        type=Path,
        default=project_root / "EyeData.xml",
        help="EyeData.xml のパス",
    )
    args = parser.parse_args()

    entries = iter_code_entries(load_root(args.xml))
    aggregated = aggregate_by_tag_and_number(entries)

    output_dir = ConfigManager().get_path("eyedata_codes_output")
    excel_path = write_excel(output_dir, aggregated)

    used = {number for _, number in aggregated}
    maximum = max(used)
    free_ranges = find_free_ranges(used, maximum)
    free_count = sum(end - start + 1 for start, end in free_ranges)

    print(f"コード一覧を {excel_path.resolve()} に出力しました")
    print(f"使用中: {len(used)} 件 (1-{maximum})  空き: {free_count} 件")
    print("空き番号:")
    print(", ".join(format_range(start, end) for start, end in free_ranges))


if __name__ == "__main__":
    main()
